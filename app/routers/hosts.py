"""
Host list management — Excel import and SSH key upload.
"""

import os
import stat
from io import BytesIO
from pathlib import Path

from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook

from app.excel_parser import ExcelParseError, parse_hosts_excel
from app.host_store import (
    clear_hosts,
    get_hosts_safe,
    remove_host,
    replace_hosts,
    resolve_host,
    update_host,
)
from app.mock_data import get_host_or_raise, is_mock_enabled, list_hosts_safe, ping_host as mock_ping_host

KEYS_DIR = Path(__file__).resolve().parent.parent.parent / "ssh_keys"

router = APIRouter(prefix="/api/hosts", tags=["hosts"])


@router.get("")
async def list_hosts():
    """Return current loaded host list (no passwords, no key paths)."""
    if is_mock_enabled():
        return {"hosts": list_hosts_safe()}
    return {"hosts": get_hosts_safe()}


@router.post("/import")
async def import_hosts(file: UploadFile = File(...)):
    """Upload Excel file to load GPU host list."""
    if is_mock_enabled():
        return {"hosts": list_hosts_safe(), "message": "Mock mode enabled: loaded built-in demo hosts."}
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=400,
            detail="File must be an Excel file (.xlsx)",
        )
    try:
        contents = await file.read()
        workbook = load_workbook(filename=BytesIO(contents), read_only=False)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {e!s}")
    try:
        hosts = parse_hosts_excel(workbook)
    except ExcelParseError as e:
        raise HTTPException(status_code=400, detail=str(e))
    if not hosts:
        raise HTTPException(
            status_code=400,
            detail="No valid rows found; ensure host_ip and username are set for at least one row.",
        )
    replace_hosts(hosts)
    return {"hosts": get_hosts_safe(), "message": f"Imported {len(hosts)} host(s)."}


@router.get("/template")
async def export_hosts_template():
    """Download an Excel template for host import."""
    wb = Workbook()
    ws = wb.active
    ws.title = "hosts"
    headers = [
        "host_ip",
        "hostname",
        "username",
        "password",
        "device_type",
        "auth_type",
        "key_path",
        "ssh_port",
        "remark",
    ]
    ws.append(headers)
    ws.append(
        [
            "10.0.0.10",
            "gpu-node-01",
            "root",
            "",
            "GPU",
            "key",
            "/home/root/.ssh/id_rsa",
            22,
            "示例：密钥认证",
        ]
    )
    ws.append(
        [
            "10.0.0.20",
            "cpu-node-01",
            "admin",
            "your_password",
            "CPU",
            "password",
            "",
            22,
            "示例：密码认证",
        ]
    )

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = "aidc_hosts_template.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.delete("/{host_id}")
async def delete_host(host_id: int):
    """Remove a single host from the loaded list."""
    if is_mock_enabled():
        _ = get_host_or_raise(host_id)
        return {"message": f"Mock mode: host {host_id} is read-only", "hosts": list_hosts_safe()}
    if not remove_host(host_id):
        raise HTTPException(status_code=404, detail=f"Host not found: {host_id}")
    return {"message": f"Host {host_id} removed", "hosts": get_hosts_safe()}


@router.delete("")
async def delete_all_hosts():
    """Clear all loaded hosts."""
    if is_mock_enabled():
        return {"message": "Mock mode: demo hosts are read-only", "hosts": list_hosts_safe()}
    count = clear_hosts()
    return {"message": f"Cleared {count} host(s)", "hosts": []}


@router.get("/{host_id}/ping")
async def ping_host(host_id: int):
    """Quick ICMP ping reachability check. Returns online status."""
    if is_mock_enabled():
        try:
            online = mock_ping_host(host_id)
        except ValueError as exc:
            raise HTTPException(status_code=404, detail=str(exc))
        return {"host_id": host_id, "online": online}
    host = resolve_host(host_id)
    if not host:
        raise HTTPException(status_code=404, detail=f"Host not found: {host_id}")
    from app.net_ping import ping_host as ping_once

    online = await ping_once(host["host_ip"], timeout_s=1.5, count=1)
    return {"host_id": host_id, "online": online}


@router.post("/{host_id}/upload-key")
async def upload_key(host_id: int, file: UploadFile = File(...)):
    """Upload an SSH private key file for a specific host.

    Saves to ssh_keys/<host_id>_<host_ip>, sets file permissions to 0600,
    and updates the host's auth_type to 'key'.
    """
    host = resolve_host(host_id)
    if not host:
        raise HTTPException(status_code=404, detail=f"Host not found: {host_id}")

    try:
        key_bytes = await file.read()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read file: {e!s}")

    if len(key_bytes) == 0:
        raise HTTPException(status_code=400, detail="Key file is empty")
    if len(key_bytes) > 64 * 1024:
        raise HTTPException(status_code=400, detail="Key file too large (max 64KB)")

    # Validate the key can be loaded by paramiko
    import tempfile
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix="_sshkey") as tmp:
            tmp.write(key_bytes)
            tmp_path = tmp.name
        from app.ssh_runner import _load_private_key
        _load_private_key(tmp_path)
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid SSH key format (supported: Ed25519, RSA, ECDSA): {e!s}",
        )
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)

    KEYS_DIR.mkdir(parents=True, exist_ok=True)
    safe_ip = host["host_ip"].replace(":", "_").replace("/", "_")
    key_filename = f"{host_id}_{safe_ip}"
    key_path = KEYS_DIR / key_filename

    with open(key_path, "wb") as f:
        f.write(key_bytes)
    os.chmod(key_path, stat.S_IRUSR | stat.S_IWUSR)  # 0600

    update_host(host_id, {"auth_type": "key", "key_path": str(key_path)})

    return {
        "message": f"Key uploaded for host {host['host_ip']}",
        "auth_type": "key",
        "host_id": host_id,
    }
